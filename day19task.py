# Create an Excel with data of Student database and add all the values which is required for student management database, Read the excel file and add the datas into a DB (using script
import openpyxl

path = "stu_1.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=5, column=3)
print(cell_obj.value)

for i in range(1, 11):
    cell_obj = sheet_obj.cell(row=5, column=i)
    print(cell_obj.value)
import mysql.connector

mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="8421436728",
)

mycursor = mydb.cursor()
print(mydb)
dbse = mydb.cursor()

dbse.execute("CREATE DATABASE Students_Management_System")
dbse = mydb.cursor()

mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="8421436728",
    database="students_management_system"
)
dbse = mydb.cursor()

dbse.execute("CREATE TABLE student3 (roll_no INT(10),Reg_no INT(10),NAME VARCHAR(255), semester1 INT(25),semester2 INT(25),semester3 INT(25), CGPA INT(35) ,PHONE_NUMBER INT,email_id VARCHAR(255))")
cur = mydb.cursor()
cur.execute('SELECT * FROM student3')
for row in cur:
    print(row)

import pandas as pd
df = pd.read_excel('stu_1.xlsx')
import xlrd
import mysql

xl_sheet = xlrd.open_workbook("stu_1.xlsx")
xl_sheet
sheet_name = xl_sheet.sheet_names()
sheet_name
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="8421436728",
    database="students_management_system"
)

cur = mydb.cursor()
for s in range(0, 1):
    sheet = xl_sheet.sheet_by_index(s)
    sql = "INSERT INTO student3(roll_no,Reg_no,NAME,semester1,semester2 ,semester3 , CGPA ,email_id) VALUES(%s,%s,%s,%s,%s,%s,%s,%s)"
    for r in range(1, sheet.nrows):
        roll_no = sheet.cell(r, 0).value
        Reg_no = sheet.cell(r, 1).value
        NAME = sheet.cell(r, 2).value
        semester1 = sheet.cell(r, 3).value
        semester2 = sheet.cell(r, 4).value
        semester3 = sheet.cell(r, 5).value
        CGPA = sheet.cell(r, 6).value
        email_id = sheet.cell(r, 7).value
        values = (roll_no, Reg_no, NAME, semester1, semester2, semester3, CGPA, email_id)

        cur.execute(sql, values)
mydb.commit()
mycursor = mydb.cursor()

mycursor.execute("SELECT * FROM student3")

myresult = mycursor.fetchall()

for x in myresult:
    print(x)
mycursor = mydb.cursor()

mycursor.execute("SELECT NAME FROM student3 WHERE CGPA >6")

myresult = mycursor.fetchall()

for x in myresult:
    print(x)
mydb.commit()

mydb.close()